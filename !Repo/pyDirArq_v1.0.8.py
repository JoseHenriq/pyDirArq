"""
ARQUIVO GERADO POR INTEGRAÇÃO AUTOMÁTICA — wrapperPandas + pyDirArq_v1.0.6
NÃO ALTERE ESTE ARQUIVO DIRETAMENTE. Modificações devem ser feitas na classe wrapperPandas.
"""

#===================================================================================================
# Classe wrapperPandas — CRUD Excel com Pandas + Openpyxl
#===================================================================================================
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional, List, Dict, Any

#--- Wrapper OOP para operações CRUD em tabela Excel usando pandas + openpyxl.
class wrapperPandas:

    EXCEL_FILE = "tabDados.xlsx"
    SHEET_NAME = "Sheet1"
    START_ROW  = 2  # linha 2 (B2)
    START_COL  = 2  # coluna B
    COLUMNS    = ["iteração", "nivel", "subnivel", "raiz", "pathAbs"]

    def __init__(self, excel_file: str = None, sheet_name: str = None):
        self.EXCEL_FILE = excel_file if excel_file else self.EXCEL_FILE
        self.SHEET_NAME = sheet_name if sheet_name else self.SHEET_NAME

    def _ensure_file_exists(self):
        path = Path(self.EXCEL_FILE)
        if not path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = self.SHEET_NAME
            
            # Cabeçalho a partir de B2
            for j, col in enumerate(self.COLUMNS, start=self.START_COL):
                cell = ws.cell(row=self.START_ROW, column=j, value=col)
                cell.font = openpyxl.styles.Font(bold=True)
                ws.column_dimensions[get_column_letter(j)].width = 18
            
            ws["A1"] = "Tabela de Dados"
            ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
            wb.save(self.EXCEL_FILE)

    def _read_table(self) -> pd.DataFrame:
        self._ensure_file_exists()
        wb = load_workbook(self.EXCEL_FILE, data_only=True)
        ws = wb[self.SHEET_NAME]
        max_row = ws.max_row
        data = []
        for row in range(self.START_ROW + 1, max_row + 1):
            row_data = []
            empty_row = True
            for col in range(self.START_COL, self.START_COL + len(self.COLUMNS)):
                value = ws.cell(row=row, column=col).value
                if value is not None:
                    empty_row = False
                row_data.append(value)
            if not empty_row:
                data.append(row_data)
        df = pd.DataFrame(data, columns=self.COLUMNS)
        return df

    def _save_table(self, df: pd.DataFrame):
        # Garante colunas na ordem correta
        if list(df.columns) != self.COLUMNS:
            df = df[self.COLUMNS]
        
        wb = load_workbook(self.EXCEL_FILE)
        ws = wb[self.SHEET_NAME]
        max_row = ws.max_row
        
        # Limpa dados anteriores
        for row in range(self.START_ROW + 1, max_row + 1):
            for col in range(self.START_COL, self.START_COL + len(self.COLUMNS)):
                ws.cell(row=row, column=col).value = None
        
        # Escreve os dados
        for i, (_, row_data) in enumerate(df.iterrows(), start=self.START_ROW + 1):
            for j, col in enumerate(self.COLUMNS, start=self.START_COL):
                ws.cell(row=i, column=j, value=row_data[col])
        
        # Remove linhas vazias extras
        current_max_row = self.START_ROW + len(df)
        for row in range(current_max_row + 1, max_row + 1):
            for col in range(self.START_COL, self.START_COL + len(self.COLUMNS)):
                ws.cell(row=row, column=col).value = None
        
        wb.save(self.EXCEL_FILE)

    def create(self, records: List[Dict[str, Any]]):
        df = self._read_table()
        new_df = pd.DataFrame(records)
        
        # Validação
        if not set(new_df.columns).issubset(set(self.COLUMNS)):
            raise ValueError(f"Colunas inválidas. Permitido: {self.COLUMNS}")
        
        for col in self.COLUMNS:
            if col not in new_df.columns:
                new_df[col] = None
        new_df = new_df[self.COLUMNS]
        
        updated_df = pd.concat([df, new_df], ignore_index=True) if len(df) > 0 else new_df
        self._save_table(updated_df)
        print(f"{len(records)} registro(s) adicionado(s).")

    def read(self, filters: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
        df = self._read_table()
        if filters:
            mask = pd.Series([True] * len(df))
            for col, val in filters.items():
                if col in self.COLUMNS:
                    mask &= (df[col] == val)
            df = df[mask].reset_index(drop=True)
        return df

    def update(self, filters: Dict[str, Any], new_values: Dict[str, Any]):
        df = self._read_table()
        mask = pd.Series([True] * len(df))
        for col, val in filters.items():
            if col in self.COLUMNS:
                mask &= (df[col] == val)
        
        if not mask.any():
            print("Nenhum registro encontrado para atualizar.")
            return
        
        for col, new_val in new_values.items():
            if col in self.COLUMNS:
                df.loc[mask, col] = new_val
        
        self._save_table(df)
        print(f"{mask.sum()} registro(s) atualizado(s).")

    def delete(self, filters: Dict[str, Any]):
        df = self._read_table()
        mask = pd.Series([True] * len(df))
        for col, val in filters.items():
            if col in self.COLUMNS:
                mask &= (df[col] == val)
        
        if not mask.any():
            print("Nenhum registro encontrado para deletar.")
            return
        
        updated_df = df[~mask].reset_index(drop=True)
        self._save_table(updated_df)
        print(f"{mask.sum()} registro(s) removido(s).")

    def print_table(self):
        df = self._read_table()
        if df.empty:
            print("Tabela vazia.")
        else:
            print(df.to_string(index=False))

#===================================================================================================
# A PARTIR DAQUI: CONTEÚDO ORIGINAL DE pyDirArq_v1.0.6.py (SEM ALTERAÇÕES)
#===================================================================================================
import os
import sys
import time
import threading
import datetime
from pathlib import Path
from tkinter import filedialog
from colorama import Back, Fore, Style, init
init()    # Initialize colorama (required for Windows)

#-------------------------------------------------------------------------------
# Instancializações e inicializações
#-------------------------------------------------------------------------------
flag_debug = True #False
BARRA_VERTICAL = 124   # |
BARRA          =  47   # /
TRACO          =  45   # -
CONTRA_BARRA   =  92   # \
    
#===================================================================================================
# Classe EstouVivo
#===================================================================================================
class EstouVivo:
    """Indicador visual 'estou vivo' para loops demorados."""

    def __init__(self, estilo='spinner'):
        self.estilo  = estilo
        self._ativo  = False
        self._thread = None
        self._inicio = None

    def _spinner(self):
        chars = [BARRA_VERTICAL, BARRA, TRACO, CONTRA_BARRA]        
        i     = 0
        
        while self._ativo:
            elapsed = time.time() - self._inicio
            sys.stdout.write(f'\r⏳ Escaneando... {chr(chars[i % 4])}  [{time.strftime("%H:%M:%S", time.gmtime(elapsed))}]')
            sys.stdout.flush()
            time.sleep(0.2)
            i += 1
        sys.stdout.write('\r' + ' ' * 50 + '\r')

        sys.stdout.flush()

    def _pontos(self):
        i = 0
        while self._ativo:
            elapsed = time.time() - self._inicio
            pontos = '.' * ((i % 6) + 1)
            sys.stdout.write(f'\r⏳ Escaneando{pontos:<6} [{elapsed:.0f}s]')
            sys.stdout.flush()
            time.sleep(0.4)
            i += 1
        sys.stdout.write('\r' + ' ' * 50 + '\r')
        sys.stdout.flush()

    def iniciar(self, mensagem='Escaneando'):
        """Ativa o indicador. Retorna o próprio objeto para uso com 'with'."""
        self._inicio = time.time()
        self._ativo  = True
        metodo       = self._spinner if self.estilo == 'spinner' else self._pontos
        self._thread = threading.Thread(target=metodo, daemon=True)
        self._thread.start()
        print(f'▶️  Estou vivo ATIVADO — {mensagem}')
        return self

    def parar(self):
        """Desativa o indicador."""
        self._ativo = False
        if self._thread:
            self._thread.join(timeout=1)
        elapsed = time.time() - self._inicio
        #print(f'⏹️  Estou vivo DESATIVADO — {elapsed:.1f}s decorridos')
        sys.stdout.write(f'⏹️  Estou vivo DESATIVADO — decorridos: [{time.strftime("%H:%M:%S", time.gmtime(elapsed))}]')

    # Suporte ao gerenciador de contexto (with)
    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.parar()

#===================================================================================================
# Funções 
#===================================================================================================
#-------------------------------------------------------------------------------
# Lê a data e hora do sistema e formata para o padrão especificado na entrada
#-------------------------------------------------------------------------------
def PreparaDataHora(formatoDataHora):
    # Dicionário para converter o dia da semana de inglês para português
    dias_da_semana = {
        'Mon': 'Seg',
        'Tue': 'Ter',
        'Wed': 'Qua',
        'Thu': 'Qui',
        'Fri': 'Sex',
        'Sat': 'Sáb',
        'Sun': 'Dom'
    }
    #---------------------------------------------------------------------
    data_hora_atual = datetime.datetime.now().strftime(formatoDataHora)
    #---------------------------------------------------------------------
    # Se o formato da data e hora contiver o dia da semana (%a), substitui o dia da semana em inglês pelo dia da semana em português
    if ('%a' in formatoDataHora) :
        dia_da_semana   = data_hora_atual[-3:]                   # extrai os últimos 3 caracteres (%a)
        dia_da_semana   = dias_da_semana [dia_da_semana]         # converte para o dia da semana em português
        data_hora_atual = data_hora_atual[:-3] + dia_da_semana   
    if ('%f' in formatoDataHora) :
        data_hora_atual = data_hora_atual[:-3]
    return data_hora_atual
#-------------------------------------------------------------------------------
# Apresentação das mensagens com estilos ajustáveis
#-------------------------------------------------------------------------------
def printStyled(text, fg=None, bg=None, bold=False, underline=False):
    """
    Prints styled colored text to the terminal.
    Args:
        text (str) : Text to print;
        fg (str)   : Foreground color: 'BLACK', 'RED', 'GREEN', 'YELLOW', 'BLUE', 'MAGENTA', 'CYAN', 'WHITE'. Default: 'WHITE'.
        bg (str)   : Background color: same options as fg. Default: None (no change).
        bold (bool): If True, text is bold.
        underline (bool): If True, text is underlined.
    """
    #--- Color maps
    # Foreground
    COLORS_FG = {
        'BLACK'  : Fore.BLACK,
        'RED'    : Fore.RED,
        'GREEN'  : Fore.GREEN,
        'YELLOW' : Fore.YELLOW,
        'BLUE'   : Fore.BLUE,
        'MAGENTA': Fore.MAGENTA,
        'CYAN'   : Fore.CYAN,
        'WHITE'  : Fore.WHITE,
    }
    # Background
    COLORS_BG = {
        'BLACK'  : Back.BLACK,
        'RED'    : Back.RED,
        'GREEN'  : Back.GREEN,
        'YELLOW' : Back.YELLOW,
        'BLUE'   : Back.BLUE,
        'MAGENTA': Back.MAGENTA,
        'CYAN'   : Back.CYAN,
        'WHITE'  : Back.WHITE,
    }
    # Defaults
    fg_code = COLORS_FG.get(fg, Fore.WHITE)
    bg_code = COLORS_BG.get(bg, '')
    style_code = ''
    if bold:
        style_code += Style.BRIGHT
    if underline:
        # ANSI underline is not well supported in all terminals, but we can try
        style_code += '\033[4m'
    # Print styled text
    #-----------------------------------------------------------------
    print(f"{bg_code}{fg_code}{style_code}{text}{Style.RESET_ALL}")
    #-----------------------------------------------------------------

#===================================================================================================
# Função principal
#===================================================================================================
def main(pathDir_, pathDirRels_):

    total_geral = 0
    intNivel    = 0
    iteracao    = 1

    flag_debug_Raiz = True
    if (flag_debug_Raiz):
        msgDebug = ""
    elif flag_debug:
        printStyled(f'\n{"TAMANHO":>12}  {"SUBDIRETÓRIO"}\n{"="*50}', fg='CYAN')
    printStyled(f'\n{"Aguarde o escaneamento ...\n"}', fg='GREEN')
    
    #--- Ativa o indicador "Estou vivo" com spinner
    with EstouVivo(estilo='spinner').iniciar('Escaneando diretório'):
        #------------------------------------------------
        for raiz, dirs, arquivos in os.walk(pathDir_):
        #------------------------------------------------
            intNivel = raiz.count('\\') - 1
            raizfrmt = " "*intNivel*4
            #print(f'{raizfrmt}{raiz}')

            if (flag_debug and not flag_debug_Raiz):
                printStyled(f'iteracao={iteracao:04d} nivel={intNivel:02d}-raiz={raizfrmt}', fg='GREEN')
                printStyled(f'-Qtidd Subdir   = {len(dirs)}', fg='YELLOW')
                print (f'dirs     = {dirs}')
                printStyled(f'-Qtidd Arquivos = {len(arquivos)}', fg='YELLOW')
                print (f'arquivos = {arquivos}\n')
            elif flag_debug_Raiz:
                msgTmp    = f'\niteracao={iteracao:04d} nivel={intNivel:02d}-raiz={raizfrmt}{raiz}'
                msgDebug += msgTmp
            iteracao += 1

            #--- TO DO: salvamento dos dados escaneados utilizando a classe wrapperPandas
            #    colunas=dado: "iteração"=iteracao; "nivel"=intNivel; "subnivel"=" "; raiz"=raiz

    #--- Finalização do escaneamento
    if (flag_debug and not flag_debug_Raiz):
        printStyled(f'{"="*50}', fg='CYAN')
        printStyled(f'{total_geral:>12,} bytes  TOTAL ({pathDir_})', fg='CYAN')
    elif flag_debug_Raiz:
        printStyled(f'\n\nRaízes a cada nível:\n{msgDebug}', fg='YELLOW')

    return

#===================================================================================================
# Ponto de entrada do programa
#===================================================================================================
if __name__ == '__main__':
    #--- Seleção pelo usuário do diretório a ser processado
    strDir  = filedialog.askdirectory(title = "Selecione o diretório a ser processado:").replace('/', '\\')
    printStyled (f"\nDiretório a ser processado: strDir = '{strDir}'", fg='GREEN')
    pathDir = Path(strDir)
    #--- Seleção pelo usuário do diretório a ser utilizado para salvamento dos relatórios
    strDirRels = filedialog.askdirectory(title = "Selecione o diretório para salvar os arquivos-relatórios:").replace('/', '\\')
    printStyled (f"Diretório para salvar os arquivos-relatórios: strDirRels = '{strDirRels}'", fg='GREEN')
    pathDirRels = Path(strDirRels)
    #--- Ativa a função Principal
    #----------------------------
    main(pathDir, pathDirRels)
    #----------------------------

#===================================================================================================
# Fim do programa
#===================================================================================================
printStyled("---> Fim do programa.", fg="GREEN")    
exit(0)

#===================================================================================================
# Fim do arquivo pyDirArq_v1.0.8.py
#===================================================================================================
