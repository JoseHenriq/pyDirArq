#===============================================================================
# Script: pandaOpenpyxl_v0101.py
# Manipulação de tabela Excel (tabDados.xlsx) com Pandas + Openpyxl
# Tabela a partir da célula B2 e colunas: iteração, nivel, subnivel, raiz, pathAbs
# Operações CRUD implementadas.
#===============================================================================
import pandas as pd
import openpyxl

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional, List, Dict, Any

EXCEL_FILE = "tabDados.xlsx"
SHEET_NAME = "Sheet1"
START_ROW  = 2  # linha 2 (B2)
START_COL  = 2  # coluna B (índice base 1 do openpyxl)
COLUMNS    = ["iteração", "nivel", "subnivel", "raiz", "pathAbs"]

#-------------------------------------------------------------------------------
# Cria o arquivo Excel com cabeçalho se não existir.
#-------------------------------------------------------------------------------
def ensure_file_exists():
    path = Path(EXCEL_FILE)
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        
        # Escreve cabeçalho a partir de B2
        for j, col in enumerate(COLUMNS, start=START_COL):
            cell = ws.cell(row=START_ROW, column=j, value=col)
            cell.font = openpyxl.styles.Font(bold=True)
            # Ajusta largura da coluna
            ws.column_dimensions[get_column_letter(j)].width = 18
        
        # Dá um título à planilha visível
        ws["A1"] = "Tabela de Dados"
        ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
        
        wb.save(EXCEL_FILE)
        print(f"Arquivo '{EXCEL_FILE}' criado com cabeçalho.")

#-------------------------------------------------------------------------------
# Lê a tabela do Excel para um DataFrame, ignorando cabeçalho e vazios.
#-------------------------------------------------------------------------------
def read_table() -> pd.DataFrame:

    ensure_file_exists()
    
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]
    
    # Encontrar o intervalo de dados (até a última linha preenchida)
    max_row = ws.max_row
    data = []
    # Começa na linha seguinte ao cabeçalho (START_ROW + 1)
    for row in range(START_ROW + 1, max_row + 1):
        row_data = []
        empty_row = True
        for col in range(START_COL, START_COL + len(COLUMNS)):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                empty_row = False
            row_data.append(value)
        # Só adiciona se tiver algum dado na linha
        if not empty_row:
            data.append(row_data)
    
    df = pd.DataFrame(data, columns=COLUMNS)
    return df

#-------------------------------------------------------------------------------
# Salva o DataFrame na tabela a partir de B2, sobrescrevendo dados existentes.
#-------------------------------------------------------------------------------
def save_table(df: pd.DataFrame):
    
    # Garante que as colunas estão na ordem correta
    if list(df.columns) != COLUMNS:
        #------------------
        df = df[COLUMNS]
        #------------------
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    # Limpa dados anteriores (a partir de START_ROW+1 até o fim)
    max_row = ws.max_row
    for row in range(START_ROW + 1, max_row + 1):
        for col in range(START_COL, START_COL + len(COLUMNS)):
            ws.cell(row=row, column=col).value = None
    
    # Escreve os dados
    for i, (_, row_data) in enumerate(df.iterrows(), start=START_ROW + 1):
        for j, col in enumerate(COLUMNS, start=START_COL):
            ws.cell(row=i, column=j, value=row_data[col])
    
    # Ajusta a última linha visível (opcional: remove linhas vazias extras)
    # Pode haver mais linhas do que precisamos após deletar
    current_max_row = START_ROW + len(df)
    for row in range(current_max_row + 1, max_row + 1):
        for col in range(START_COL, START_COL + len(COLUMNS)):
            ws.cell(row=row, column=col).value = None
    
    wb.save(EXCEL_FILE)

#-------------------------------------------------------------------------------
# Adiciona novos registros à tabela.
#-------------------------------------------------------------------------------
def create(records: List[Dict[str, Any]]):
    df     = read_table()
    new_df = pd.DataFrame(records)
    
    # Valida colunas
    if not set(new_df.columns).issubset(set(COLUMNS)):
        raise ValueError(f"Colunas inválidas. Permitido: {COLUMNS}")
    
    # Garante todas as colunas presentes
    for col in COLUMNS:
        if col not in new_df.columns:
            new_df[col] = None
    new_df = new_df[COLUMNS]
    
    if len(df) == 0:
        updated_df = new_df
    else:
        updated_df = pd.concat([df, new_df], ignore_index=True)
    
    #------------------------
    save_table(updated_df)
    #------------------------
    print(f"{len(records)} registro(s) adicionado(s).")

#-------------------------------------------------------------------------------
# Lê registros, com filtro opcional por coluna=valor.
#-------------------------------------------------------------------------------
def read(filters: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
    df = read_table()
    if filters:
        mask = pd.Series([True] * len(df))
        for col, val in filters.items():
            if col in COLUMNS:
                mask &= (df[col] == val)
        df = df[mask].reset_index(drop=True)
    return df

#-------------------------------------------------------------------------------
# Atualiza registros que correspondam ao filtro com novos valores.
#-------------------------------------------------------------------------------
def update(filters: Dict[str, Any], new_values: Dict[str, Any]):
    
    df = read_table()
    
    # Aplica filtro
    mask = pd.Series([True] * len(df))
    for col, val in filters.items():
        if col in COLUMNS:
            mask &= (df[col] == val)
    
    if not mask.any():
        print("Nenhum registro encontrado para atualizar.")
        return
    
    # Atualiza os valores
    for col, new_val in new_values.items():
        if col in COLUMNS:
            df.loc[mask, col] = new_val
    
    save_table(df)
    print(f"{mask.sum()} registro(s) atualizado(s).")

#-------------------------------------------------------------------------------
# Remove registros que correspondem ao filtro.
#-------------------------------------------------------------------------------
def delete(filters: Dict[str, Any]):
    df = read_table()
    
    mask = pd.Series([True] * len(df))
    for col, val in filters.items():
        if col in COLUMNS:
            mask &= (df[col] == val)
    
    if not mask.any():
        print("Nenhum registro encontrado para deletar.")
        return
    
    # Mantém apenas os que NÃO correspondem ao filtro
    updated_df = df[~mask].reset_index(drop=True)
    save_table(updated_df)
    print(f"{mask.sum()} registro(s) removido(s).")

#-------------------------------------------------------------------------------
# Exibe a tabela atual no console.
#-------------------------------------------------------------------------------
def print_table():
    #-------------------
    df = read_table()
    #-------------------
    
    if df.empty:
        print("Tabela vazia.")
    else:
        print(df.to_string(index=False))

#===============================================================================
# Ponto de entrada no programa
#===============================================================================
if __name__ == "__main__":

    # Exemplo de uso das operações CRUD
    print("--- Teste CRUD com tabDados.xlsx ---\n")

    #--- CREATE
    registros = [
        {"iteração": 1, "nivel": "A", "subnivel": "A1", "raiz": "C:", "pathAbs": "C:\\projetos\\foo"},
        {"iteração": 2, "nivel": "B", "subnivel": "B1", "raiz": "D:", "pathAbs": "D:\\dados\\bar"},
        {"iteração": 3, "nivel": "A", "subnivel": "A2", "raiz": "C:", "pathAbs": "C:\\projetos\\baz"},
    ]
    #-------------------
    create(registros)
    #-------------------
    
    print("\n--- Tabela após CREATE ---")
    print_table()

    #--- READ com filtro
    print("\n--- READ - registros com nivel='A' ---")
    #------------------------------------------
    df_filtro = read(filters={"nivel": "A"})
    #------------------------------------------
    print(df_filtro.to_string(index=False))

    #--- UPDATE
    print("\n--- UPDATE - alterar subnivel='A1' para subnivel='AX' ---")
    #-----------------------------------------------------------------------------------
    update(filters={"subnivel": "A1"}, new_values={"subnivel": "AX", "iteração": 99})
    #-----------------------------------------------------------------------------------
    print_table()

    #--- DELETE
    print("\n--- DELETE - remover registros com nivel='B' ---")
    #--------------------------------
    delete(filters={"nivel": "B"})
    #--------------------------------
    print_table()

    #--- Finaliza o programa
    print("\n--- Operações CRUD concluídas. Verifique o arquivo 'tabDados.xlsx'. ---")
