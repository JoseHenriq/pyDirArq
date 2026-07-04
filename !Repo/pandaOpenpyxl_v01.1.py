import pandas as pd
from openpyxl import load_workbook

# ===========================================================================
# Função 1: Classificar registros por coluna (ascendente)
# ===========================================================================
def classificar_por_coluna(workbook, sheet, intervalo, coluna_ordenar):
    """
    Classifica os registros de uma planilha Excel por uma coluna específica (ascendente).
    
    Args:
        workbook (str)      : Caminho do arquivo .xlsx
        sheet (str)         : Nome da planilha
        intervalo (str)     : Intervalo de células (ex: 'B4:F20')
        coluna_ordenar (str): Nome da coluna para classificar
    
    Retorna:
        pd.DataFrame: DataFrame com os dados classificados
    """
    # Lê o intervalo especificado
    df = pd.read_excel(workbook, sheet_name=sheet, header=None)
    
    # Pula as 3 primeiras linhas (cabeçalho) e pega as linhas de dados
    # Como o intervalo começa em B4, as 3 primeiras linhas são cabeçalho
    df_dados = df.iloc[3:19, 1:6]  # B=1, C=2, D=3, E=4, F=5 (0-indexed)
    
    # Define os nomes das colunas
    df_dados.columns = ['iteração', 'nivel', 'subnivel', 'raiz', 'pathRelativo']
    
    # Remove linhas totalmente vazias
    df_dados = df_dados.dropna(how='all')
    
    # Converte a coluna 'subnivel' para numérico (se possível)
    df_dados['subnivel'] = pd.to_numeric(df_dados['subnivel'], errors='coerce')
    
    # Classifica pela coluna especificada (ascendente)
    df_ordenado = df_dados.sort_values(by=coluna_ordenar, ascending=True)
    
    # Reseta o índice
    df_ordenado = df_ordenado.reset_index(drop=True)
    
    print(f'Registros classificados por "{coluna_ordenar}" (ascendente):')
    print(f'Total de registros: {len(df_ordenado)}')
    print(df_ordenado.to_string(index=False))
    
    return df_ordenado

# ===========================================================================
# Função 2: Criar nova sheet no workbook
# ===========================================================================
def criar_nova_sheet(workbook, nome_nova_sheet, df=None):
    """
    Cria uma nova planilha no workbook especificado.
    Opcionalmente, insere dados de um DataFrame.
    
    Args:
        workbook (str): Caminho do arquivo .xlsx
        nome_nova_sheet (str): Nome da nova planilha
        df (pd.DataFrame, opcional): Dados para inserir na nova sheet
    
    Retorna:
        bool: True se criou com sucesso, False caso contrário
    """
    try:
        # Carrega o workbook existente
        wb = load_workbook(workbook)
        
        # Verifica se a sheet já existe
        if nome_nova_sheet in wb.sheetnames:
            print(f'Sheet "{nome_nova_sheet}" já existe no workbook.')
            return False
        
        # Cria a nova sheet
        wb.create_sheet(title=nome_nova_sheet)
        
        # Se um DataFrame foi fornecido, insere os dados
        if df is not None:
            ws = wb[nome_nova_sheet]
            
            # Insere o cabeçalho
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Insere os dados
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx + 2, column=col_idx, value=value)
        
        # Salva o workbook
        wb.save(workbook)
        print(f'Sheet "{nome_nova_sheet}" criada com sucesso em "{workbook}".')
        return True
        
    except FileNotFoundError:
        print(f'Workbook "{workbook}" não encontrado.')
        return False
    except Exception as e:
        print(f'Erro ao criar sheet: {e}')
        return False

# ===========================================================================
# Exemplo de uso
# ===========================================================================
if __name__ == '__main__':
    WORKBOOK = 'tabPyDirArq_001.xlsx'
    SHEET    = 'Sheet1'
    INTERVALO = 'B4:F20'
    COLUNA_ORDENAR = 'subnivel'
    
    # 1) Classifica os dados
    df_classificado = classificar_por_coluna(WORKBOOK, SHEET, INTERVALO, COLUNA_ORDENAR)
    
    # 2) Cria uma nova sheet com os dados classificados
    criar_nova_sheet(WORKBOOK, 'Classificado', df_classificado)
    
    # 3) Cria uma nova sheet vazia
    criar_nova_sheet(WORKBOOK, 'Resumo')

