#===================================================================================================
# Programa : pyDirArq_v1.1.1.py
# Descrição: salva arquivos e subdiretórios de um diretorio totalizando o tamanho dos arquivos e
# subdiretórios.
#---------------------------------------------------------------------------------------------------
# Data : 28/06/2026 dom
#---------------------------------------------------------------------------------------------------
# Utilização: >python pyDirArq_v1.1.0.py
# --------------------------------------------------------------------------------------------------
# Versão 1.1.1 - organização dos dados da tabela;
# Versão 1.1.0 - recodificação da função wrapper _save_table(self, df: pd.DataFrame) para utilização
#                da função "escrever_linha(self, ws, i, row_data)" para escrever 0 na tabela;
#                ---> problema na extensão "XLSX, CSV, TSV & Markdown Editor - VS Code Extension"
#                do VSCode qto a apresentação de 0s.
# Versão 1.0.9 - integração: wrapperPandas + salvamento dos dados do os.walk em tabPyDirArq_001.xlsx
# Versão 1.0.8 - implementa a classe wrapperPandas;
# Versão 1.0.7 - implementa wrapper CRUD do Pandas e Openpyxl;
# Versão 1.0.6 - correção do nível para qtidd de '\' no path;
# Versão 1.0.5 - adicionado ao script 1.0.4 "EstouVivo";
# Versão 1.0.4 - apresenta a raíz cada iteração for: 'for raiz, dirs, arquivos in os.walk(pathDir_)'
# Versão 1.0.3 - apresenta os subdiretórios e arquivos de todos os níveis;
# Versão 1.0.2 - apresenta os subdiretórios e arquivos do Nível1;
# Versão 1.0.1 - apresenta os subdiretórios e arquivos;
# Versão 1.0.0 - versão inicial;
#===================================================================================================

#===================================================================================================
# Imports e inicializações
#===================================================================================================
import os
import sys
import time
import threading
import datetime
from pathlib import Path
from tkinter import filedialog

from colorama import Back, Fore, Style, init
# Instancia a classe colorama
init()     

flag_debug     = True

BARRA_VERTICAL = 124
BARRA          =  47
TRACO          =  45
CONTRA_BARRA   =  92

nomeArquivoExcel = 'tabPyDirArq_001.xlsx'
nomePlanilha     = 'Sheet1'

#===================================================================================================
# Classe wrapperPandas — CRUD Excel com Pandas + Openpyxl
#===================================================================================================
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional, List, Dict, Any

# Wrapper OOP para operações CRUD em tabela Excel usando pandas + openpyxl.
class wrapperPandas:

    #--- Valores default
    EXCEL_FILE = "tabDados.xlsx"  # será sobrescrito no construtor
    SHEET_NAME = "Sheet1"
    START_ROW  = 3
    START_COL  = 2
    COLUMNS    = ["iteração", "nivel", "subnivel", "raiz", "pathRelativo"]

    #--- Inicializações na instancialização de objeto wrapperPandas
    def __init__(self, excel_file: str = None, sheet_name: str = None):
        self.EXCEL_FILE = excel_file if excel_file else self.EXCEL_FILE
        self.SHEET_NAME = sheet_name if sheet_name else self.SHEET_NAME

    #--- Se não existe o arquivo Excel, cria-o.
    def _ensure_file_exists(self):
        path = Path(self.EXCEL_FILE)
        if not path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = self.SHEET_NAME
            for j, col in enumerate(self.COLUMNS, start=self.START_COL):
                cell      = ws.cell(row=self.START_ROW, column=j, value=col)
                cell.font = openpyxl.styles.Font(bold=True)
                ws.column_dimensions[get_column_letter(j)].width = 18
            ws["B2"]      = "Tabela de Dados"
            ws["B2"].font = openpyxl.styles.Font(bold=True, size=14)
            wb.save(self.EXCEL_FILE)

    #--- Le tabela no arquivo Excel
    def _read_table(self) -> pd.DataFrame:
        self._ensure_file_exists()
        wb      = load_workbook(self.EXCEL_FILE, data_only=True)
        ws      = wb[self.SHEET_NAME]
        max_row = ws.max_row
        data    = []
        for row in range(self.START_ROW + 1, max_row + 1):
            row_data  = []
            empty_row = True
            for col in range(self.START_COL, self.START_COL + len(self.COLUMNS)):
                value = ws.cell(row=row, column=col).value
                if value is not None:
                    empty_row = False
                row_data.append(value)
            if not empty_row:
                data.append(row_data)
        df = pd.DataFrame(data, columns=self.COLUMNS)
        return df

    #--- Cria tabela
    def create(self, records: List[Dict[str, Any]]):
        df     = self._read_table()
        new_df = pd.DataFrame(records)
        if not set(new_df.columns).issubset(set(self.COLUMNS)):
            raise ValueError(f"Colunas inválidas. Permitido: {self.COLUMNS}")
        for col in self.COLUMNS:
            if col not in new_df.columns:
                new_df[col] = None
        new_df     = new_df[self.COLUMNS]
        updated_df = pd.concat([df, new_df], ignore_index=True) if len(df) > 0 else new_df
        self._save_table(updated_df)

    #--- Salva tabela
    '''
    def _save_table_new(self, df: pd.DataFrame):

        #--- Atualiza a planilha Excel com, eventualmente, novo df
        # 1- colunas
        if list(df.columns) != self.COLUMNS:
            df = df[self.COLUMNS]
        #-------------------------------------
        wb = load_workbook(self.EXCEL_FILE)
        #-------------------------------------
        # 2- linhas
        ws      = wb[self.SHEET_NAME]
        max_row = ws.max_row
        # 3- limpa os conteúdos da planilha atual
        for row in range(self.START_ROW + 1, max_row + 1):
            for col in range(self.START_COL, self.START_COL + len(self.COLUMNS)):
                ws.cell(row=row, column=col).value = None
        # 4- armazena os dados atuais na planilha
        for i, (_, row_data) in enumerate(df.iterrows(), start=self.START_ROW + 1):

            #--- DEBUG: mostra o que está vindo do DataFrame
            print(f'Linha {i}:')
            for col in self.COLUMNS:
                valor_bruto = row_data[col]  # acesso direto, sem .get()
                print(f'  coluna "{col}" = {repr(valor_bruto)} (tipo: {type(valor_bruto).__name__})')

            for j, col in enumerate(self.COLUMNS, start=self.START_COL):

                # Pega o valor bruto
                valor = row_data[col]
                
                # Trata explicitamente: se for NaN ou None, vira 0
                if valor is None or (isinstance(valor, float) and pd.isna(valor)):
                    valor = 0
                
                # Garante que 0 seja inteiro, não float
                if isinstance(valor, float) and valor == int(valor):
                    valor = int(valor)
                
                ws.cell(row=i, column=j, value=valor)

        #--- Salva a planilha Excel atualizada no arquivo (workbook) Excel
        wb.save(self.EXCEL_FILE)
    '''
    def _save_table(self, df: pd.DataFrame):

        print(df)

        # Classifica pela coluna especificada (ascendente)
        #df_ordenado = df.sort_values(by='subnivel', ascending=True)
        df = df.sort_values(by='subnivel', ascending=True)
        # Reseta o índice
        #df_ordenado = df_ordenado.reset_index(drop=True)
        df = df.reset_index(drop=True)
    
        #--- Atualiza a planilha Excel com, eventualmente, novo df
        if list(df.columns) != self.COLUMNS:
            df = df[self.COLUMNS]
        
        # FORÇA colunas numéricas para int (resolve o problema do 0)
        colunas_int = ['iteração', 'nivel', 'subnivel']
        for col in colunas_int:
            if col in df.columns:
                df[col] = df[col].fillna(0).astype(int)
        
        wb = load_workbook(self.EXCEL_FILE)
        ws = wb[self.SHEET_NAME]
        max_row = ws.max_row
        
        # Limpa os conteúdos da planilha atual
        for row in range(self.START_ROW + 1, max_row + 1):
            for col in range(self.START_COL, self.START_COL + len(self.COLUMNS)):
                ws.cell(row=row, column=col).value = None
        
        # Armazena os dados atuais na planilha
        for i, (_, row_data) in enumerate(df.iterrows(), start=self.START_ROW + 1):
            for j, col in enumerate(self.COLUMNS, start=self.START_COL):
                valor = row_data[col]
                # Converte numpy types para Python nativos
                if hasattr(valor, 'item'):
                    valor = valor.item()
                ws.cell(row=i, column=j, value=valor)
        
        wb.save(self.EXCEL_FILE)

    #--- Lê tabela
    def read(self, filters: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
        df = self._read_table()
        if filters:
            mask = pd.Series([True] * len(df))
            for col, val in filters.items():
                if col in self.COLUMNS:
                    mask &= (df[col] == val)
            df = df[mask].reset_index(drop=True)
        return df

    #--- Atualiza tabela
    def update(self, filters: Dict[str, Any], new_values: Dict[str, Any]):
        df   = self._read_table()
        mask = pd.Series([True] * len(df))
        for col, val in filters.items():
            if col in self.COLUMNS:
                mask &= (df[col] == val)
        if not mask.any():
            print("Nenhum registro encontrado para atualizar.")
            return
        for col, new_val in new_values.items():
            if col in self.COLUMNS:
                df.loc[mask, col] = new_val
        self._save_table(df)

    #--- Deleta registro
    def delete(self, filters: Dict[str, Any]):
        df   = self._read_table()
        mask = pd.Series([True] * len(df))
        for col, val in filters.items():
            if col in self.COLUMNS:
                mask &= (df[col] == val)
        if not mask.any():
            print("Nenhum registro encontrado para deletar.")
            return
        updated_df = df[~mask].reset_index(drop=True)
        self._save_table(updated_df)

    #--- Apresenta a tabela no console.
    def print_table(self):
        df = self._read_table()
        if df.empty:
            print("Tabela vazia.")
        else:
            print(df.to_string(index=False))

#===================================================================================================
# Classe EstouVivo
#===================================================================================================
class EstouVivo:

    def __init__(self, estilo='spinner'):
        self.estilo  = estilo
        self._ativo  = False
        self._thread = None
        self._inicio = None

    def _spinner(self):
        chars = [BARRA_VERTICAL, BARRA, TRACO, CONTRA_BARRA]
        i     = 0
        while self._ativo:
            elapsed = time.time() - self._inicio
            sys.stdout.write(f'\r⏳ Escaneando... {chr(chars[i % 4])}  [{time.strftime("%H:%M:%S", time.gmtime(elapsed))}]')
            sys.stdout.flush()
            time.sleep(0.2)
            i += 1
        sys.stdout.write('\r' + ' ' * 50 + '\r')
        sys.stdout.flush()

    def _pontos(self):
        i = 0
        while self._ativo:
            elapsed = time.time() - self._inicio
            pontos = '.' * ((i % 6) + 1)
            sys.stdout.write(f'\r⏳ Escaneando{pontos:<6} [{elapsed:.0f}s]')
            sys.stdout.flush()
            time.sleep(0.4)
            i += 1
        sys.stdout.write('\r' + ' ' * 50 + '\r')
        sys.stdout.flush()

    def iniciar(self, mensagem='Escaneando'):
        self._inicio = time.time()
        self._ativo  = True
        metodo       = self._spinner if self.estilo == 'spinner' else self._pontos
        self._thread = threading.Thread(target=metodo, daemon=True)
        self._thread.start()
        print(f'▶️  Estou vivo ATIVADO — {mensagem}')
        return self

    def parar(self):
        self._ativo = False
        if self._thread:
            self._thread.join(timeout=1)
        elapsed = time.time() - self._inicio
        sys.stdout.write(f'⏹️  Estou vivo DESATIVADO — decorridos: [{time.strftime("%H:%M:%S", time.gmtime(elapsed))}]')

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.parar()

#===================================================================================================
# Funções auxiliares
#===================================================================================================
def PreparaDataHora(formatoDataHora):
    dias_da_semana = {'Mon': 'Seg','Tue': 'Ter','Wed': 'Qua','Thu': 'Qui','Fri': 'Sex','Sat': 'Sáb','Sun': 'Dom'}
    data_hora_atual = datetime.datetime.now().strftime(formatoDataHora)
    if ('%a' in formatoDataHora):
        dia_da_semana = data_hora_atual[-3:]
        dia_da_semana = dias_da_semana[dia_da_semana]
        data_hora_atual = data_hora_atual[:-3] + dia_da_semana
    if ('%f' in formatoDataHora):
        data_hora_atual = data_hora_atual[:-3]
    return data_hora_atual

def printStyled(text, fg=None, bg=None, bold=False, underline=False):
    COLORS_FG = {'BLACK': Fore.BLACK,'RED': Fore.RED,'GREEN': Fore.GREEN,'YELLOW': Fore.YELLOW,'BLUE': Fore.BLUE,'MAGENTA': Fore.MAGENTA,'CYAN': Fore.CYAN,'WHITE': Fore.WHITE}
    COLORS_BG = {'BLACK': Back.BLACK,'RED': Back.RED,'GREEN': Back.GREEN,'YELLOW': Back.YELLOW,'BLUE': Back.BLUE,'MAGENTA': Back.MAGENTA,'CYAN': Back.CYAN,'WHITE': Back.WHITE}
    fg_code = COLORS_FG.get(fg, Fore.WHITE)
    bg_code = COLORS_BG.get(bg, '')
    style_code = Style.BRIGHT if bold else ''
    if underline:
        style_code += '\033[4m'
    print(f"{bg_code}{fg_code}{style_code}{text}{Style.RESET_ALL}")

import pandas as pd
from openpyxl import load_workbook

#===================================================================================================
# Função principal COM INTEGRAÇÃO wrapperPandas
#===================================================================================================
def main(pathDir_, pathDirRels_):

    total_geral = 0
    intNivel    = 0
    iteracao    = 1

    flag_debug_Raiz = True
    if (flag_debug_Raiz):
        msgDebug = ""
    elif flag_debug:
        printStyled(f'\n{"TAMANHO":>12}  {"SUBDIRETÓRIO"}\n{"="*50}', fg='CYAN')
    printStyled(f'\n{"Aguarde o escaneamento ...\n"}', fg='GREEN')
    
    # --- INSTÂNCIA DO WRAPPER para o arquivo tabPyDirArq_001.xlsx ---
    #--------------------------------------------------------------------------
    wp = wrapperPandas(excel_file=nomeArquivoExcel, sheet_name=nomePlanilha)
    #--------------------------------------------------------------------------
    
    #--- Ativa o indicador "Estou vivo" com spinner
    with EstouVivo(estilo='spinner').iniciar('Escaneando diretório'):

        #------------------------------------------------
        for raiz, dirs, arquivos in os.walk(pathDir_):
        #------------------------------------------------
            intNivel     = raiz.count('\\')     # - 1
            if iteracao == 1 : subNivelBase = intNivel
            subnivel     = int(intNivel - subNivelBase)
            pathRelativo = os.path.relpath(raiz, pathDir)

            raizfrmt = " "*intNivel*4

            if (flag_debug and not flag_debug_Raiz):
                printStyled(f'iteracao={iteracao:04d} nivel={intNivel:02d}-raiz={raizfrmt}', fg='GREEN')
                printStyled(f'-Qtidd Subdir   = {len(dirs)}', fg='YELLOW')
                print (f'dirs     = {dirs}')
                printStyled(f'-Qtidd Arquivos = {len(arquivos)}', fg='YELLOW')
                print (f'arquivos = {arquivos}\n')
            elif flag_debug_Raiz:
                msgTmp       = f'\niteracao={iteracao:04d} nivel={intNivel:02d}-raiz={raizfrmt}{raiz}'
                msgDebug    += msgTmp
                
            #--- ✅ INTEGRAÇÃO: salvamento dos dados escaneados utilizando wrapperPandas ---
            registro = {
                "iteração"     : iteracao,
                "nivel"        : intNivel,
                "subnivel"     : subnivel,
                "raiz"         : raiz,
                "pathRelativo" : pathRelativo
            }
            #-----------------------
            wp.create([registro])
            #-----------------------

            #--- (Opcional: para melhor performance, coletar em lista e salvar ao final) ---

            iteracao += 1

    #--- Finalização do escaneamento ---
    if (flag_debug and not flag_debug_Raiz):
        printStyled(f'{"="*50}', fg='CYAN')
        printStyled(f'{total_geral:>12,} bytes  TOTAL ({pathDir_})', fg='CYAN')
    elif flag_debug_Raiz:
        printStyled(f'\n\nRaízes a cada nível:\n{msgDebug}', fg='YELLOW')

    #--- Informação adicional sobre salvamento ---
    printStyled(f"\n✅ Dados salvos em {nomeArquivoExcel}/{nomePlanilha} (tabela a partir de B2)", fg='GREEN')

    return

#===================================================================================================
# Ponto de entrada do programa
#===================================================================================================
if __name__ == '__main__':

    # Diretório a ser processado
    strDir  = filedialog.askdirectory(title = "Selecione o diretório a ser processado:").replace('/', '\\')
    printStyled (f"\nDiretório a ser processado: strDir = '{strDir}'", fg='GREEN')
    pathDir = Path(strDir)
    
    # Diretório para salvar os arquivos-relatórios
    strDirRels = filedialog.askdirectory(title = "Selecione o diretório para salvar os arquivos-relatórios:").replace('/', '\\')
    printStyled (f"Diretório para salvar os arquivos-relatórios: strDirRels = '{strDirRels}'", fg='GREEN')
    pathDirRels = Path(strDirRels)
    
    #----------------------------
    main(pathDir, pathDirRels)
    #----------------------------

#===================================================================================================
# Fim do programa
#===================================================================================================
printStyled("---> Fim do programa.", fg="GREEN")
exit(0)

#===================================================================================================
# Fim do arquivo pyDirArq_v1.1.1.py
#===================================================================================================
